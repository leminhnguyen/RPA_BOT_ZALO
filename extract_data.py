from ZaloFunctions import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors, Color


class ExtractData():
    FIELDS = [
        "Mã đại lý",
        "Họ và tên",
        "Tổng Phí Phải Thu",
        "Tổng Phí Thực Thu",
        "Còn PT",
        "Tỷ lệ"
    ]
    COLUMNS = {
        "Phải thu": "AI",
        "Thực thu BK06": "AJ",
        "Nhóm": "I",
        "Mã đại lý": "J",
        "Họ và tên": "K"
    }

    FEE_SHEET = "TDTP"
    FIRST_ROW: int = 5  # first row of the table
    FEE_SUMMARY: str = "fee_summary.xlsx"
    FEE_TEMPLATE = CURRENT_DIRECTORY + "/templates/fee_template.xlsx"

    AGG_CORLOR = "FFC7CE"  # color in the agg row of summary sheet

    def __init__(self):
        self.file_path = ""
        self.wb = None
        self.ws = None
        self.data_to_send = None
        self.fee_summary = ""
        self.data_status = False
        self.excel = Dispatch("Excel.Application")

    def _hide_excel(self):
        self.excel.DisplayAlerts = False
        self.excel.visible = False
        return


    def _show_excel(self):
        self.excel.DisplayAlerts  = True
        self.excel.visible = True
        return


    def _copy_fee_data(self):
        """
        Copy value of fee sheet to template file
        """
        wb = openpyxl.load_workbook(self.FEE_TEMPLATE)
        save_excel(self.fee_summary, wb)
        #copy fee data to template
        fee_summary = self.excel.Workbooks.Open(self.fee_summary)
        fee_data    = self.excel.Workbooks.Open(self.file_path)
        fee_data.Worksheets(self.FEE_SHEET).Cells.Copy()
        fee_summary.Worksheets(self.FEE_SHEET).Cells.PasteSpecial(Paste=-4163)
        fee_data.Close(False)
        fee_summary.Close(True)

    def browse_file(self):
        path = get_excel_file(title="Chọn file tổng hợp thu phí")
        if path == "":
            return
        else:
            self.file_path = path

    def validate_fee_data(self):
        self._hide_excel()
        if self.file_path == "":
            messagebox.showerror(
                "Lỗi",
                "Chưa chọn file thông tin thu phí!"
            )
            return
        wb = self.excel.Workbooks.Open(self.file_path)
        #check if fee_sheet exists
        try:
            wb.Worksheets(self.FEE_SHEET)
        except:
            messagebox.showerror(
                "Lỗi",
                "Không tìm thấy sheet " + self.FEE_SHEET + " \n" + "Xin vui lòng kiểm tra lại!"
            )
            return
        wb.Close(False)
        fee_summary = self.file_path.split("/")[:-1]
        fee_summary.append(self.FEE_SUMMARY)
        self.fee_summary = '/'.join(fee_summary)
        self._copy_fee_data()
        self.wb = openpyxl.load_workbook(self.FEE_SUMMARY)
        self.ws = self.wb[self.FEE_SHEET]
        self._show_excel()

    @staticmethod
    def str_to_number(st):
        if st == "" or st is None:
            return 0
        else:
            return float(st)

    def _get_aggregate(self):
        """
        return dictonary describing Group - Agent name - fee
        {group_name:
            {[agent_id * agent_name]:
                {tong phai thu: [],
                tong thuc thu: []
                }
            }
        }
        """
        last_row = v_empty_cell(self.ws, self.COLUMNS["Nhóm"])
        summary = dict()
        for i in range(self.FIRST_ROW + 1, last_row, 1):
            group_name = self.ws[self.COLUMNS["Nhóm"] + str(i)].value
            agent_id = self.ws[self.COLUMNS["Mã đại lý"] + str(i)].value
            agent_name = self.ws[self.COLUMNS["Họ và tên"] + str(i)].value
            phai_thu = self.ws[self.COLUMNS["Phải thu"] + str(i)].value
            thuc_thu = self.ws[self.COLUMNS["Thực thu BK06"] + str(i)].value
            agent = agent_id + "*" + agent_name
            if group_name not in summary.keys():
                summary.update({group_name: dict()})
            if agent not in summary[group_name].keys():
                summary[group_name].update(
                    {agent: {"Phải thu": [], "Thực thu": []}})
            summary[group_name][agent]["Phải thu"].append(self.str_to_number(phai_thu))
            summary[group_name][agent]["Thực thu"].append(self.str_to_number(thuc_thu))
        return summary

    def _get_groups(self):
        """
        :return: dictionary of agent sets; {agent: set({agent_name:, agent_id})}
        """
        d = dict()
        last_row = v_empty_cell(self.ws, self.COLUMNS["Nhóm"])
        for i in range(self.FIRST_ROW + 1, last_row, 1):
            group_name = self.ws[self.COLUMNS["Nhóm"] + str(i)].value
            agent_id = self.ws[self.COLUMNS["Mã đại lý"] + str(i)].value
            agent_name = self.ws[self.COLUMNS["Họ và tên"] + str(i)].value
            agent = {"Mã đại lý": agent_id, "Họ và tên": agent_name}
            if group_name not in d.keys():
                d.update({group_name: []})
            if agent not in d[group_name]:
                d[group_name].append(agent)
        return d

    def _create_data_to_send(self):
        """
        create data to send in the form of {group_name: []}
        """
        if self.ws is None:
            return
        tables = dict()
        summary = self._get_aggregate()
        for group in summary.keys():
            d = summary[group]
            tables.update({group: []})
            for agent in d.keys():
                agent_id = agent.split("*")[0]
                agent_name = agent.split("*")[1]
                phai_thu = sum(d[agent]["Phải thu"])
                thuc_thu = sum(d[agent]["Thực thu"])
                con_pt = phai_thu - thuc_thu
                if thuc_thu == 0:
                    ty_le = 0
                else:
                    ty_le = phai_thu / thuc_thu
                tables[group].append([agent_id, agent_name, phai_thu, thuc_thu, con_pt, ty_le])
        self.data_to_send = tables

    def create_worksheets_to_send(self):
        """
        combining data to worksheets to send
        """
        if self.ws is None:
            return
        self._create_data_to_send()
        phai_thu_col = self.FIELDS.index("Tổng Phí Phải Thu") + 1
        thuc_thu_col = self.FIELDS.index("Tổng Phí Thực Thu") + 1
        ty_le_col = self.FIELDS.index("Tỷ lệ") + 1
        con_pt_col = self.FIELDS.index("Còn PT") + 1
        wb = openpyxl.load_workbook(self.FEE_SUMMARY)
        for group, table in self.data_to_send.items():
            wb.copy_worksheet(wb["template"])
            ws = wb["template Copy"]
            ws.title = group
            # fill agent summary
            for row, data in enumerate(table):
                tong_phai_thu = 0
                tong_thuc_thu = 0
                for col, value in enumerate(data):
                    ws.cell(row + self.FIRST_ROW + 1, col + 1).value = value
                    if col == phai_thu_col:
                        tong_phai_thu += value
                    if col == thuc_thu_col:
                        tong_thuc_thu += value
            # summary row
            agg_row = self.FIRST_ROW + 1 + len(table)
            ws.cell(agg_row, 1).value = "Tổng"
            ws.cell(agg_row, phai_thu_col).value = tong_phai_thu
            ws.cell(agg_row, thuc_thu_col).value = tong_thuc_thu
            ws.cell(agg_row, con_pt_col).value = \
                tong_thuc_thu - tong_thuc_thu
            if tong_phai_thu == 0:
                ty_le = 0
            else:
                ty_le = tong_thuc_thu / tong_phai_thu
            ws.cell(agg_row, ty_le_col).value = \
                ty_le
            # format
            for col in range(1, len(self.FIELDS) + 1, 1):
                ws.cell(agg_row, col).font = Font(bold=True)
                ws.cell(agg_row, col).fill =PatternFill(
                    start_color=self.AGG_CORLOR,
                    end_color=self.AGG_CORLOR,
                    fill_type='solid'
                )
        save_excel(self.FEE_SUMMARY, wb)

if __name__ == '__main__':
    x = ExtractData()
    x.browse_file()
    x.validate_fee_data()
    start = time.time()
    x.create_worksheets_to_send()
    print(time.time() -start)
