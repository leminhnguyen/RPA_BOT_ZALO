# _*_ coding: utf-8
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.webdriver import FirefoxProfile
from selenium.webdriver.common.action_chains import ActionChains

import time
import os
import openpyxl
import keyboard
from tempfile import TemporaryDirectory

import tkinter as tk
from tkinter import Tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import Button
from tkinter import Frame
from tkinter import PhotoImage
from tkinter import Label

from PIL import ImageTk,Image

from win32com.client import Dispatch
import win32com.client
import shutil

from ctypes import windll


CURRENT_DIRECTORY = str(os.getcwd()).replace("\\","/")


def get_excel_file(title = "Excel files", directory = CURRENT_DIRECTORY):
    """
    locate folder containning excel file
    """
    filename = ''
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    filename = filedialog.askopenfilename(
        initialdir = directory,
        title = title,
        filetypes = [('Excel files', '*.xlsx')]
    )
    return filename

def message(title, info):
    """
    show message
    """
##    root = Tk()
##    root.withdraw()
##    root.attributes("-topmost", True)
    messagebox.showinfo(title, info)

def error(title, info):

    """
    show message
    """
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showerror(title, info)

def v_empty_cell(ws, col):
    """
    return the flast empty cell in a column of a worksheet
    """
    row = ws.max_row
    while ws[col + str(row)].value in [None,""] and row > 0:
        row -= 1
    return row+1

def h_empty_cell(ws, row):
    """
    return the last empty cell in a row of a worksheet
    """
    col = ws.max_column
    while ws.cell(row = row, column = col).value in [None, ""] and column > 0:
        col -= 1
    return chr(ord("A") + col)

def next_char(c):
    """
    return next char
    """
    return str(chr(ord(c)+1))

def pre_char(c):
    """
    return previous char
    """
    return str(chr(ord(c)-1))

def clear_clipboard():
    if windll.user32.OpenClipboard(None):
        windll.user32.EmptyClipboard()
        windll.user32.CloseClipboard()


def normalize(st):
    """
    Delete all space in string
    """
    return str(st).replace(" ","")

def save_excel(excel_file, wb):
    """
    show popup if excel file is being opened
    """
    while True:
        try:
            wb.save(excel_file)
            break
        except PermissionError:
            Tk().withdraw()
            messagebox.showerror('Lỗi', 'Hãy đóng file ' + excel_file)
    wb.save(excel_file)

if __name__ == '__main__':
    print (CURRENT_DIRECTORY)
